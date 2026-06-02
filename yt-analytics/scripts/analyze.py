#!/usr/bin/env python3
"""
Analyze YouTube comments and generate Excel workbook with insights.

Uses Claude API (free via Anthropic) for intelligent categorization,
or falls back to rule-based analysis if no API key is set.

Setup:
  pip install anthropic openpyxl pandas

Usage:
  export ANTHROPIC_API_KEY=sk-ant-...   # optional but better analysis
  python scripts/analyze.py
  python scripts/analyze.py --input outputs/comments_raw.csv
"""

import csv
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("pip install pandas", file=sys.stderr)
    sys.exit(1)

SCRIPTS_DIR = Path(__file__).parent
ROOT_DIR = SCRIPTS_DIR.parent
OUTPUT_DIR = ROOT_DIR / "outputs"

# ── Colors ─────────────────────────────────────────────────────────────────
RED     = "D62828"
GOLD    = "F4C430"
DARK    = "1A1A1A"
MID     = "2D2D2D"
GREEN   = "2ECC71"
BLUE    = "3498DB"
GRAY    = "888888"
WHITE   = "FFFFFF"
LIGHT   = "F5F5F5"


def load_comments(path: Path) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def ai_analyze(comments: list[dict]) -> dict:
    """Use Claude to find patterns — falls back to rule-based if no key."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("  (no ANTHROPIC_API_KEY — using rule-based analysis)")
        return rule_based_analyze(comments)

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        # Sample up to 300 comments for analysis (token budget)
        sample = comments[:300]
        texts = [f"- {c['text'][:200]}" for c in sample]
        batch = "\n".join(texts)

        prompt = f"""Analyze these {len(sample)} YouTube comments and return a JSON object with:
{{
  "top_questions": ["list of the 10 most common questions asked"],
  "content_requests": ["list of the 10 most requested video topics"],
  "top_tools_mentioned": {{"tool_name": count}},
  "sentiment_breakdown": {{"positive": N, "neutral": N, "negative": N, "question": N}},
  "key_themes": ["5 recurring themes in the comments"],
  "high_priority_replies": ["5 comments that most deserve a creator reply — quote them"],
  "creator_insights": ["3 actionable insights for the creator based on these comments"]
}}

Comments:
{batch}

Return only valid JSON, no other text."""

        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        # Strip markdown code fences if present
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        return json.loads(raw)

    except Exception as e:
        print(f"  AI analysis failed ({e}), using rule-based")
        return rule_based_analyze(comments)


def rule_based_analyze(comments: list[dict]) -> dict:
    """Fallback analysis without AI."""
    tool_keywords = [
        "claude", "chatgpt", "gpt", "gemini", "cursor", "copilot",
        "airtable", "notion", "zapier", "make", "n8n", "vercel",
        "python", "javascript", "api", "github", "codex",
    ]
    tools = Counter()
    questions = []
    requests = []
    themes = Counter()

    for c in comments:
        text = c["text"].lower()
        # Tools
        for t in tool_keywords:
            if t in text:
                tools[t] += 1
        # Questions
        if "?" in c["text"] and len(c["text"]) > 20:
            questions.append(c["text"][:120])
        # Content requests
        if any(w in text for w in ["please make", "can you make", "video on", "tutorial on", "cover ", "do a video"]):
            requests.append(c["text"][:120])
        # Themes
        if any(w in text for w in ["beginner", "getting started", "how to start"]):
            themes["Beginner questions"] += 1
        if any(w in text for w in ["free", "cost", "price", "pay", "subscription"]):
            themes["Pricing / cost"] += 1
        if any(w in text for w in ["api", "code", "script", "automat"]):
            themes["Technical / coding"] += 1
        if any(w in text for w in ["workflow", "system", "process", "productivity"]):
            themes["Workflows & productivity"] += 1
        if any(w in text for w in ["thank", "love", "amazing", "great video"]):
            themes["Positive feedback"] += 1

    sentiments = {"positive": 0, "neutral": 0, "negative": 0, "question": 0}
    for c in comments:
        cat = c.get("category", "general")
        if cat == "positive_feedback":
            sentiments["positive"] += 1
        elif cat == "question":
            sentiments["question"] += 1
        elif cat == "issue_report":
            sentiments["negative"] += 1
        else:
            sentiments["neutral"] += 1

    top_q = list(dict.fromkeys(questions))[:10]
    top_r = list(dict.fromkeys(requests))[:10]

    return {
        "top_questions": top_q if top_q else ["No clear questions found — try pulling more comments"],
        "content_requests": top_r if top_r else ["No explicit requests found"],
        "top_tools_mentioned": dict(tools.most_common(10)),
        "sentiment_breakdown": sentiments,
        "key_themes": [k for k, _ in themes.most_common(5)] or ["General engagement"],
        "high_priority_replies": [c["text"][:150] for c in sorted(comments, key=lambda x: int(x.get("likes", 0)), reverse=True)[:5]],
        "creator_insights": [
            "Most engaged comments are questions — consider a FAQ video",
            "High tool-mention rate suggests technical audience",
            "Reply to top-liked comments to boost engagement signals",
        ],
    }


def cell_style(ws, cell_ref, bold=False, bg=None, fg=WHITE, size=11, align="left", wrap=False):
    cell = ws[cell_ref] if isinstance(cell_ref, str) else cell_ref
    cell.font = Font(bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell


def header_row(ws, row, cols: list, bg=DARK, fg=WHITE):
    for col, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=label)
        cell_style(ws, c, bold=True, bg=bg, fg=fg, align="center")


def make_workbook(comments: list[dict], analysis: dict, out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── TAB 1: OVERVIEW ──────────────────────────────────────────────────
    ws_ov = wb.create_sheet("📊 Overview")
    ws_ov.sheet_view.showGridLines = False
    ws_ov.column_dimensions["A"].width = 28
    ws_ov.column_dimensions["B"].width = 18
    ws_ov.column_dimensions["C"].width = 42

    # Title
    ws_ov.merge_cells("A1:C1")
    c = ws_ov["A1"]
    c.value = "YouTube Comment Intelligence Report"
    cell_style(ws_ov, c, bold=True, bg=RED, fg=WHITE, size=16, align="center")
    ws_ov.row_dimensions[1].height = 36

    ws_ov.merge_cells("A2:C2")
    c = ws_ov["A2"]
    c.value = f"Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  ·  {len(comments)} total comments"
    cell_style(ws_ov, c, bg=DARK, fg=GRAY, align="center")

    # Stats
    videos = list({c["video_title"] for c in comments})
    total_likes = sum(int(c.get("likes", 0)) for c in comments)
    q_count = sum(1 for c in comments if c.get("category") == "question")
    req_count = sum(1 for c in comments if c.get("category") == "content_request")

    stats = [
        ("Total Comments", len(comments), ""),
        ("Videos Analyzed", len(videos), ""),
        ("Total Likes on Comments", total_likes, ""),
        ("Questions Detected", q_count, f"{q_count/max(len(comments),1)*100:.0f}% of comments"),
        ("Content Requests", req_count, f"{req_count/max(len(comments),1)*100:.0f}% of comments"),
    ]

    ws_ov.row_dimensions[3].height = 8
    header_row(ws_ov, 4, ["Metric", "Value", "Notes"])
    for i, (label, val, note) in enumerate(stats, 5):
        ws_ov.cell(i, 1, label)
        ws_ov.cell(i, 2, val)
        ws_ov.cell(i, 3, note)
        bg = LIGHT if i % 2 == 0 else WHITE
        for col in range(1, 4):
            c = ws_ov.cell(i, col)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws_ov.cell(i, 2).alignment = Alignment(horizontal="center", vertical="center")

    # Sentiment breakdown
    ws_ov.row_dimensions[10].height = 8
    sent = analysis.get("sentiment_breakdown", {})
    ws_ov.merge_cells("A11:C11")
    c = ws_ov["A11"]
    c.value = "Sentiment Breakdown"
    cell_style(ws_ov, c, bold=True, bg=MID, fg=WHITE, align="center")

    header_row(ws_ov, 12, ["Category", "Count", "% of Total"], bg=DARK)
    sent_colors = {"positive": GREEN, "question": BLUE, "neutral": GRAY, "negative": RED}
    for i, (k, v) in enumerate(sent.items(), 13):
        pct = f"{v/max(len(comments),1)*100:.1f}%"
        ws_ov.cell(i, 1, k.title())
        ws_ov.cell(i, 2, v)
        ws_ov.cell(i, 3, pct)
        col = sent_colors.get(k, GRAY)
        ws_ov.cell(i, 1).fill = PatternFill("solid", fgColor=col)
        ws_ov.cell(i, 1).font = Font(color=WHITE, bold=True)
        ws_ov.cell(i, 2).alignment = Alignment(horizontal="center", vertical="center")

    # Pie chart for sentiment
    pie = PieChart()
    pie.title = "Comment Sentiment Mix"
    labels = Reference(ws_ov, min_col=1, min_row=13, max_row=12 + len(sent))
    data = Reference(ws_ov, min_col=2, min_row=12, max_row=12 + len(sent))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.style = 10
    pie.width = 14
    pie.height = 10
    ws_ov.add_chart(pie, "E4")

    # ── TAB 2: INSIGHTS ──────────────────────────────────────────────────
    ws_in = wb.create_sheet("💡 Insights")
    ws_in.sheet_view.showGridLines = False
    ws_in.column_dimensions["A"].width = 6
    ws_in.column_dimensions["B"].width = 60
    ws_in.column_dimensions["C"].width = 20

    ws_in.merge_cells("A1:C1")
    c = ws_in["A1"]
    c.value = "Creator Insights & Recommendations"
    cell_style(ws_in, c, bold=True, bg=RED, fg=WHITE, size=14, align="center")
    ws_in.row_dimensions[1].height = 32

    sections = [
        ("🎯 Actionable Creator Insights", analysis.get("creator_insights", []), GREEN),
        ("❓ Top Questions from Audience", analysis.get("top_questions", []), BLUE),
        ("🎬 Requested Video Topics", analysis.get("content_requests", []), GOLD),
        ("🔑 Key Themes", analysis.get("key_themes", []), MID),
    ]

    row = 3
    for title, items, color in sections:
        ws_in.row_dimensions[row].height = 6
        row += 1
        ws_in.merge_cells(f"A{row}:C{row}")
        c = ws_in[f"A{row}"]
        c.value = title
        cell_style(ws_in, c, bold=True, bg=color, fg=WHITE if color != GOLD else DARK, size=12, align="left")
        ws_in.row_dimensions[row].height = 26
        row += 1

        for i, item in enumerate(items or ["(none found)"]):
            ws_in.cell(row, 1, i + 1)
            ws_in.merge_cells(f"B{row}:C{row}")
            c = ws_in.cell(row, 2, item)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            bg = LIGHT if i % 2 == 0 else WHITE
            for col in [1, 2]:
                ws_in.cell(row, col).fill = PatternFill("solid", fgColor=bg)
            ws_in.row_dimensions[row].height = 40
            row += 1
        row += 1

    # ── TAB 3: TOOLS MENTIONED ───────────────────────────────────────────
    ws_tools = wb.create_sheet("🔧 Tools Mentioned")
    ws_tools.sheet_view.showGridLines = False
    ws_tools.column_dimensions["A"].width = 28
    ws_tools.column_dimensions["B"].width = 16
    ws_tools.column_dimensions["C"].width = 24

    ws_tools.merge_cells("A1:C1")
    c = ws_tools["A1"]
    c.value = "Tools & Technologies Mentioned in Comments"
    cell_style(ws_tools, c, bold=True, bg=DARK, fg=WHITE, size=14, align="center")
    ws_tools.row_dimensions[1].height = 32

    tools = analysis.get("top_tools_mentioned", {})
    header_row(ws_tools, 3, ["Tool / Technology", "Mentions", "% of Comments"], bg=MID)
    total_tool_mentions = sum(tools.values()) or 1
    for i, (tool, count) in enumerate(sorted(tools.items(), key=lambda x: -x[1]), 4):
        pct = f"{count/len(comments)*100:.1f}%"
        ws_tools.cell(i, 1, tool.title())
        ws_tools.cell(i, 2, count)
        ws_tools.cell(i, 3, pct)
        bg = LIGHT if i % 2 == 0 else WHITE
        for col in range(1, 4):
            ws_tools.cell(i, col).fill = PatternFill("solid", fgColor=bg)
        ws_tools.cell(i, 2).alignment = Alignment(horizontal="center")

    # Bar chart
    if tools:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Tool Mentions"
        bar.y_axis.title = "Mentions"
        bar.style = 11
        bar.width = 20
        bar.height = 12
        data = Reference(ws_tools, min_col=2, min_row=3, max_row=3 + len(tools))
        cats = Reference(ws_tools, min_col=1, min_row=4, max_row=3 + len(tools))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws_tools.add_chart(bar, "E3")

    # ── TAB 4: HIGH PRIORITY REPLIES ─────────────────────────────────────
    ws_rep = wb.create_sheet("💬 Reply Queue")
    ws_rep.sheet_view.showGridLines = False
    ws_rep.column_dimensions["A"].width = 4
    ws_rep.column_dimensions["B"].width = 20
    ws_rep.column_dimensions["C"].width = 55
    ws_rep.column_dimensions["D"].width = 10
    ws_rep.column_dimensions["E"].width = 40

    ws_rep.merge_cells("A1:E1")
    c = ws_rep["A1"]
    c.value = "High-Priority Reply Queue"
    cell_style(ws_rep, c, bold=True, bg=RED, fg=WHITE, size=14, align="center")
    ws_rep.row_dimensions[1].height = 32

    ws_rep.merge_cells("A2:E2")
    ws_rep["A2"].value = "These comments most deserve a reply — sorted by likes + engagement signal"
    ws_rep["A2"].font = Font(color=GRAY, italic=True)

    header_row(ws_rep, 4, ["#", "Author", "Comment", "Likes", "Link"], bg=DARK)

    # Sort by likes
    sorted_comments = sorted(comments, key=lambda x: int(x.get("likes", 0)), reverse=True)
    for i, c_data in enumerate(sorted_comments[:50], 5):
        ws_rep.cell(i, 1, i - 4)
        ws_rep.cell(i, 2, c_data.get("author", ""))
        ws_rep.cell(i, 3, c_data.get("text", "")[:280])
        ws_rep.cell(i, 4, c_data.get("likes", 0))
        link_cell = ws_rep.cell(i, 5, "Open on YouTube →")
        link_cell.hyperlink = c_data.get("url", "")
        link_cell.font = Font(color="1155CC", underline="single")
        bg = LIGHT if i % 2 == 0 else WHITE
        for col in range(1, 5):
            ws_rep.cell(i, col).fill = PatternFill("solid", fgColor=bg)
        ws_rep.cell(i, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws_rep.row_dimensions[i].height = 50
        ws_rep.cell(i, 4).alignment = Alignment(horizontal="center")

    # ── TAB 5: RAW COMMENTS ──────────────────────────────────────────────
    ws_raw = wb.create_sheet("📝 Raw Comments")
    ws_raw.sheet_view.showGridLines = False
    fields = ["video_title", "author", "text", "likes", "reply_count", "category", "published_at", "url"]
    col_widths = [35, 20, 60, 10, 14, 18, 20, 42]

    header_row(ws_raw, 1, [f.replace("_", " ").title() for f in fields], bg=DARK)
    for col, w in enumerate(col_widths, 1):
        ws_raw.column_dimensions[get_column_letter(col)].width = w

    for i, c_data in enumerate(comments, 2):
        for col, field in enumerate(fields, 1):
            val = c_data.get(field, "")
            if field in ("likes", "reply_count"):
                val = int(val) if val else 0
            ws_raw.cell(i, col, val)
        ws_raw.cell(i, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws_raw.row_dimensions[i].height = 45
        bg = LIGHT if i % 2 == 0 else WHITE
        for col in range(1, len(fields) + 1):
            if col != len(fields):  # skip URL col (has hyperlink)
                ws_raw.cell(i, col).fill = PatternFill("solid", fgColor=bg)

    wb.save(out_path)
    print(f"\n✓ Workbook saved: {out_path}")
    print(f"  Tabs: Overview · Insights · Tools · Reply Queue · Raw Comments ({len(comments)} rows)")


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=str(OUTPUT_DIR / "comments_raw.csv"))
    parser.add_argument("--output", default=str(OUTPUT_DIR / "yt_comment_insights.xlsx"))
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: {input_path} not found. Run pull_comments.py first.", file=sys.stderr)
        sys.exit(1)

    comments = load_comments(input_path)
    print(f"Loaded {len(comments)} comments from {input_path.name}")

    print("Running analysis...")
    analysis = ai_analyze(comments)

    make_workbook(comments, analysis, Path(args.output))


if __name__ == "__main__":
    main()
